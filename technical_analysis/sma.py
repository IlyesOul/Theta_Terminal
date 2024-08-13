import yfinance as yf
from openpyxl import Workbook, load_workbook
import xlsxwriter
from alpha_vantage.techindicators import TechIndicators
import csv


Key = 'EJJLLU3V2H84RZPO'


techInd = TechIndicators(key = Key, output_format = 'pandas')

sma = techInd.get_sma('AAPL', time_period=50)

sma.to_csv("sma_data.csv")

print(sma)




#workbook = xlsxwriter.Workbook("DataExcel.xlsx")
#worksheet = workbook.add_worksheet("firstsheet")
#aapl = yf.Ticker("AAPL")
#print(aapl.quarterly_balancesheet)
#balance_sheet_df = aapl.quarterly_balancesheet
#balance_sheet_df.to_csv("DataExcel.xlsx")
#book = load_workbook('data.xlsx')
#sheet = book.active
#worksheet.write('A1', 'Mohamed')
#workbook.close()